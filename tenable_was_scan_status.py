"""
Tenable Vulnerability Management — WAS Scan Status Fetcher
Produces a two-sheet Excel workbook:
  Sheet 1 — Executive Summary  (KPI tiles + status breakdown table)
  Sheet 2 — Full Extract       (all scan details, filterable)

Usage:
    Ensure a .env file exists in the same directory with:
        TENABLE_ACCESS_KEY=your_key
        TENABLE_SECRET_KEY=your_secret
    Then run:
        python tenable_was_scan_status.py

Output: was_scan_status.xlsx  (written next to this script)
Requires: pip install openpyxl requests python-dotenv
"""

import os
import sys
import time
import requests
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
_SCRIPT_DIR = Path(__file__).parent

_env_path = _SCRIPT_DIR / ".env"
if not _env_path.exists():
    print(f"[ERROR] .env file not found at: {_env_path}")
    sys.exit(1)
load_dotenv(dotenv_path=_env_path)

ACCESS_KEY  = os.environ.get("TENABLE_ACCESS_KEY", "")
SECRET_KEY  = os.environ.get("TENABLE_SECRET_KEY", "")
OUTPUT_FILE = _SCRIPT_DIR / "was_scan_status.xlsx"  # always written next to the script

if not ACCESS_KEY or not SECRET_KEY:
    print("[ERROR] Missing API credentials.")
    print("  Ensure your .env file contains:")
    print("    TENABLE_ACCESS_KEY=your_access_key")
    print("    TENABLE_SECRET_KEY=your_secret_key")
    sys.exit(1)

BASE_URL       = "https://cloud.tenable.com"
HEADERS        = {
    "X-ApiKeys"   : f"accessKey={ACCESS_KEY}; secretKey={SECRET_KEY}",
    "Accept"      : "application/json",
    "Content-Type": "application/json",
}
RETRY_ATTEMPTS = 3      # retries on 429 / 5xx
RETRY_BACKOFF  = 10     # seconds to wait before first retry (doubles each attempt)

# ─────────────────────────────────────────────────────────────────────────────
# HTTP HELPER
# ─────────────────────────────────────────────────────────────────────────────
def post(endpoint: str, payload: dict = None) -> dict:
    """
    POST with retry logic for 429 / 5xx and consistent error handling.
    Never logs response bodies to avoid leaking sensitive data.
    """
    url   = f"{BASE_URL}{endpoint}"
    delay = RETRY_BACKOFF

    for attempt in range(1, RETRY_ATTEMPTS + 1):
        try:
            r = requests.post(url, headers=HEADERS, json=payload or {}, timeout=30)
        except requests.exceptions.SSLError:
            print(f"[ERROR] SSL verification failed for {endpoint}. Aborting.")
            sys.exit(1)
        except requests.exceptions.ConnectionError:
            print(f"[ERROR] Could not reach {BASE_URL}. Check network connectivity.")
            sys.exit(1)
        except requests.exceptions.Timeout:
            print(f"[WARN] Request timed out for {endpoint} (attempt {attempt}/{RETRY_ATTEMPTS}).")
            if attempt < RETRY_ATTEMPTS:
                time.sleep(delay)
                delay *= 2
                continue
            return {}

        if r.status_code == 401:
            print("[ERROR] Authentication failed (HTTP 401). Verify your API keys.")
            sys.exit(1)
        if r.status_code == 403:
            print(f"[WARN] Access denied (HTTP 403) for {endpoint}. Check API key permissions.")
            return {}
        if r.status_code == 429:
            retry_after = int(r.headers.get("Retry-After", delay))
            print(f"[WARN] Rate limited (HTTP 429). Waiting {retry_after}s before retry "
                  f"(attempt {attempt}/{RETRY_ATTEMPTS})...")
            time.sleep(retry_after)
            delay *= 2
            continue
        if r.status_code >= 500:
            print(f"[WARN] Server error HTTP {r.status_code} for {endpoint} "
                  f"(attempt {attempt}/{RETRY_ATTEMPTS}).")
            if attempt < RETRY_ATTEMPTS:
                time.sleep(delay)
                delay *= 2
                continue
            return {}
        if not r.ok:
            print(f"[WARN] POST {endpoint} returned HTTP {r.status_code}. Skipping.")
            return {}

        try:
            return r.json()
        except ValueError:
            # Catches json.JSONDecodeError on all requests versions (pre-2.28 and post)
            print(f"[WARN] POST {endpoint} returned a non-JSON response. Skipping.")
            return {}

    return {}

# ─────────────────────────────────────────────────────────────────────────────
# TENABLE WAS API
# ─────────────────────────────────────────────────────────────────────────────
def fetch_all_was_configs() -> list:
    """
    POST /was/v2/configs/search
    Paginates using 'offset' and 'limit' in the request body.
    Response: { "items": [...], "pagination": { "total": N, "offset": N, "limit": N } }
    """
    all_configs, offset, limit = [], 0, 100
    print("[*] Fetching WAS scan configurations...")

    while True:
        data  = post("/was/v2/configs/search", {"offset": offset, "limit": limit})
        items = data.get("items", [])
        total = data.get("pagination", {}).get("total", len(all_configs) + len(items))

        if not items:
            break

        all_configs.extend(items)
        print(f"    {len(all_configs)} / {total} fetched")

        if len(all_configs) >= total or len(items) < limit:
            break
        offset += limit

    return all_configs


def fetch_last_scan(config_id: str) -> dict:
    """
    POST /was/v2/scans/search
    Returns the most recent scan execution for the given config_id.
    """
    if not config_id:
        return {}

    payload = {
        "filter": {
            "and": [{"field": "config_id", "operator": "eq", "value": config_id}]
        },
        "sort"  : [{"field": "started_at", "order": "desc"}],
        "limit" : 1,
        "offset": 0,
    }
    data  = post("/was/v2/scans/search", payload)
    items = data.get("items", [])
    return items[0] if items else {}

# ─────────────────────────────────────────────────────────────────────────────
# UTILITIES
# ─────────────────────────────────────────────────────────────────────────────
def fmt_ts(ts) -> str:
    """Convert epoch seconds or ISO 8601 string to a readable UTC timestamp."""
    if not ts:
        return "N/A"
    fmt = "%Y-%m-%d %H:%M:%S UTC"
    try:
        if isinstance(ts, (int, float)):
            return datetime.fromtimestamp(ts, tz=timezone.utc).strftime(fmt)
        # ISO 8601 string — e.g. "2025-01-15T10:30:00.000Z" or "2025-01-15T10:30:00+00:00"
        # Replace trailing Z only (Python < 3.11 fromisoformat does not accept Z suffix)
        ts_str = ts if not str(ts).endswith("Z") else str(ts)[:-1] + "+00:00"
        return datetime.fromisoformat(ts_str).astimezone(timezone.utc).strftime(fmt)
    except Exception:
        return str(ts)  # return raw value rather than crash


def extract_urls(config: dict) -> str:
    """Extract target URLs from a WAS scan config's settings block."""
    s    = config.get("settings", {})
    urls = s.get("urls") or s.get("target_urls") or [s.get("url") or s.get("start_url")]
    flat = []
    for item in (urls or []):
        if isinstance(item, dict):
            flat.append(item.get("url", ""))
        elif item:
            flat.append(str(item))
    return "\n".join(filter(None, flat)) or "N/A"


def normalise_status(status: str) -> str:
    """Map raw API status values to consistent display labels."""
    mapping = {
        ""         : "Unknown",
        "completed": "Completed",
        "running"  : "Running",
        "error"    : "Error",
        "aborted"  : "Aborted",
        "never_run": "Never Run",
        "queued"   : "Queued",
        "paused"   : "Paused",
        "stopping" : "Stopping",
    }
    return mapping.get(status.lower().replace(" ", "_"), status.title() or "Unknown")

# ─────────────────────────────────────────────────────────────────────────────
# STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
NAVY       = "1F4E79"
DARK_NAVY  = "17375E"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "D6E4F0"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F2F2"
MID_GREY   = "BFBFBF"
DARK_GREY  = "595959"

STATUS_BG = {
    "Completed": "C6EFCE",
    "Running"  : "FFEB9C",
    "Error"    : "FFC7CE",
    "Aborted"  : "FFC7CE",
    "Never Run": "EDEDED",
    "Queued"   : "DDEBF7",
    "Paused"   : "FCE4D6",
    "Stopping" : "FFEB9C",
    "Unknown"  : "EDEDED",
}
STATUS_FG = {
    "Completed": "276221",
    "Running"  : "7D6608",
    "Error"    : "9C0006",
    "Aborted"  : "9C0006",
    "Never Run": "595959",
    "Queued"   : "1F4E79",
    "Paused"   : "843C0C",
    "Stopping" : "7D6608",
    "Unknown"  : "595959",
}

# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def fill(colour: str) -> PatternFill:
    return PatternFill("solid", start_color=colour, fgColor=colour)

def border(colour: str = MID_GREY) -> Border:
    s = Side(style="thin", color=colour)
    return Border(left=s, right=s, top=s, bottom=s)

def bottom_border(colour: str = MID_GREY) -> Border:
    return Border(bottom=Side(style="thin", color=colour))

def font(bold=False, size=10, colour=None, name="Arial", italic=False) -> Font:
    return Font(name=name, bold=bold, size=size, color=colour or "000000", italic=italic)

def align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 1 — EXECUTIVE SUMMARY
# ─────────────────────────────────────────────────────────────────────────────
EXTRACT_SHEET = "WAS Full Extract"

def build_summary_sheet(ws, rows: list, generated_at: str) -> None:
    ws.sheet_view.showGridLines = False

    # Column layout: A = left margin, B–F = content, G = right margin
    for col, width in [("A", 3), ("B", 24), ("C", 18), ("D", 18), ("E", 18), ("F", 18), ("G", 3)]:
        ws.column_dimensions[col].width = width

    # ── Banner ────────────────────────────────────────────────────────────────
    for row, height in [(1, 8), (2, 45), (3, 20), (4, 12)]:
        ws.row_dimensions[row].height = height

    ws.merge_cells("B2:F2")
    ws["B2"].value     = "WAS Scan Status — Executive Summary"
    ws["B2"].font      = font(bold=True, size=18, colour=WHITE)
    ws["B2"].fill      = fill(NAVY)
    ws["B2"].alignment = align("left", "center")

    ws.merge_cells("B3:F3")
    ws["B3"].value     = f"Tenable Vulnerability Management  ·  Generated: {generated_at}"
    ws["B3"].font      = font(size=10, colour=WHITE, italic=True)
    ws["B3"].fill      = fill(DARK_NAVY)
    ws["B3"].alignment = align("left", "center")

    for row_num, bg in [(2, NAVY), (3, DARK_NAVY)]:
        for col in ("A", "G"):
            ws[f"{col}{row_num}"].fill = fill(bg)

    # ── KPI Tiles (rows 5–7) ──────────────────────────────────────────────────
    for row, height in [(5, 14), (6, 36), (7, 22), (8, 12)]:
        ws.row_dimensions[row].height = height

    n             = len(rows)
    status_counts = Counter(r["Status"] for r in rows)

    tiles = [
        ("Total Scans",   n,                                               MID_BLUE,  WHITE),
        ("Completed",     status_counts.get("Completed", 0),               "217346",  WHITE),
        ("Running",       status_counts.get("Running", 0),                 "BF8F00",  "3D3D3D"),
        ("Error / Abort", status_counts.get("Error", 0) +
                          status_counts.get("Aborted", 0),                 "C00000",  WHITE),
        ("Never Run",     status_counts.get("Never Run", 0),               DARK_GREY, WHITE),
    ]

    for col_letter, (label, value, bg, fg) in zip(["B", "C", "D", "E", "F"], tiles):
        for row_num, val, bold, size in [(5, label, False, 9), (6, value, True, 22), (7, "", False, 9)]:
            c           = ws[f"{col_letter}{row_num}"]
            c.value     = val
            c.font      = font(bold=bold, size=size, colour=fg)
            c.fill      = fill(bg)
            c.alignment = align("center", "center" if row_num == 6 else "bottom")

    # ── Status Breakdown Table (rows 9+) ──────────────────────────────────────
    for row, height in [(9, 20), (10, 18)]:
        ws.row_dimensions[row].height = height

    ws.merge_cells("B9:F9")
    ws["B9"].value     = "STATUS BREAKDOWN"
    ws["B9"].font      = font(bold=True, size=9, colour=MID_BLUE)
    ws["B9"].alignment = align("left", "bottom")
    ws["B9"].border    = bottom_border(MID_BLUE)

    for col_l, hdr in zip(["B", "C", "D", "E"], ["Status", "Count", "% of Total", "Health"]):
        c           = ws[f"{col_l}10"]
        c.value     = hdr
        c.font      = font(bold=True, size=10, colour=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = align("center", "center")
        c.border    = border(NAVY)

    status_rows = [
        ("Completed", "Healthy ✅"),
        ("Running",   "In Progress 🔄"),
        ("Error",     "Needs Attention ❌"),
        ("Aborted",   "Needs Attention ❌"),
        ("Never Run", "Not Started ⬜"),
        ("Queued",    "Pending ⏳"),
        ("Paused",    "Paused ⏸"),
        ("Stopping",  "Stopping 🔄"),
        ("Unknown",   "Unknown ❓"),
    ]

    for i, (status_label, health) in enumerate(status_rows):
        row_num = 11 + i
        count   = status_counts.get(status_label, 0)
        pct     = count / n if n > 0 else 0
        bg_c    = STATUS_BG.get(status_label, WHITE)
        fg_c    = STATUS_FG.get(status_label, "000000")
        row_bg  = LIGHT_GREY if i % 2 == 0 else WHITE

        ws.row_dimensions[row_num].height = 18

        cell_data = {
            "B": (status_label,              align("left",   "center"), font(bold=True, size=10, colour=fg_c), fill(bg_c)),
            "C": (count,                     align("center", "center"), font(size=10),                         fill(row_bg)),
            "D": (pct,                       align("center", "center"), font(size=10),                         fill(row_bg)),
            "E": (health if count > 0 else "—",
                                             align("center", "center"), font(size=10, colour=fg_c if count > 0 else MID_GREY),
                                                                                                                fill(row_bg)),
        }
        for col_l, (val, aln, fnt, fll) in cell_data.items():
            c           = ws[f"{col_l}{row_num}"]
            c.value     = val
            c.font      = fnt
            c.fill      = fll
            c.alignment = aln
            c.border    = border()
            if col_l == "D":
                c.number_format = "0.0%"

    # ── Footer ────────────────────────────────────────────────────────────────
    footer_row = 11 + len(status_rows) + 2
    ws.row_dimensions[footer_row].height = 14
    ws.merge_cells(f"B{footer_row}:F{footer_row}")
    fc           = ws[f"B{footer_row}"]
    fc.value     = f"Full scan details available in '{EXTRACT_SHEET}' sheet  ·  Source: Tenable.io WAS v2 API"
    fc.font      = font(size=8, colour=MID_GREY, italic=True)
    fc.alignment = align("left", "center")

# ─────────────────────────────────────────────────────────────────────────────
# SHEET 2 — FULL EXTRACT
# ─────────────────────────────────────────────────────────────────────────────
EXTRACT_COLS = [
    ("Scan Name",      40),
    ("Last Run (UTC)", 22),
    ("Status",         14),
    ("Notes",          45),
    ("URLs",           50),
    ("Config ID",      38),
]

def build_extract_sheet(ws, rows: list) -> None:
    ws.sheet_view.showGridLines = True

    # Header row
    ws.row_dimensions[1].height = 28
    for col_idx, (col_name, width) in enumerate(EXTRACT_COLS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
        c           = ws.cell(row=1, column=col_idx, value=col_name)
        c.font      = font(bold=True, size=11, colour=WHITE)
        c.fill      = fill(NAVY)
        c.alignment = align("center", "center", wrap=True)
        c.border    = border(DARK_NAVY)

    # Data rows
    for row_idx, row in enumerate(rows, start=2):
        row_bg = LIGHT_BLUE if row_idx % 2 == 0 else WHITE
        status = row.get("Status", "")
        ws.row_dimensions[row_idx].height = 18

        for col_idx, (col_name, _) in enumerate(EXTRACT_COLS, start=1):
            c           = ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border    = border()

            if col_name == "Status":
                c.font = font(size=10, bold=True, colour=STATUS_FG.get(status, "000000"))
                c.fill = fill(STATUS_BG.get(status, WHITE))
            else:
                c.font = font(size=10)
                c.fill = fill(row_bg)

    # Freeze header + auto-filter
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(EXTRACT_COLS))}1"

# ─────────────────────────────────────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────────────────────────────────────
def main():
    configs = fetch_all_was_configs()
    if not configs:
        print("[!] No WAS scan configurations returned. Verify your API keys and WAS licence.")
        sys.exit(0)

    print(f"\n[*] Processing {len(configs)} configs...\n")
    rows = []

    for config in configs:
        config_id = config.get("id", "")
        name      = config.get("name", "Unknown")
        notes     = config.get("description", config.get("notes", ""))
        urls      = extract_urls(config)

        last_scan = fetch_last_scan(config_id)
        if last_scan:
            status   = normalise_status(last_scan.get("status", ""))
            last_run = fmt_ts(last_scan.get("started_at") or last_scan.get("finalized_at"))
        else:
            status   = "Never Run"
            last_run = "N/A"

        rows.append({
            "Scan Name"     : name,
            "Last Run (UTC)": last_run,
            "Status"        : status,
            "Notes"         : notes,
            "URLs"          : urls,
            "Config ID"     : config_id,
        })

        icon = {"Completed": "✅", "Running": "🔄", "Error": "❌", "Aborted": "❌"}.get(status, "⬜")
        print(f"  {icon} {name}  [{status}]  {last_run}")

    generated_at = datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    wb           = Workbook()

    ws_summary       = wb.active
    ws_summary.title = "Executive Summary"
    build_summary_sheet(ws_summary, rows, generated_at)

    ws_extract = wb.create_sheet(EXTRACT_SHEET)
    build_extract_sheet(ws_extract, rows)

    wb.save(OUTPUT_FILE)
    print(f"\n[✓] Exported {len(rows)} scans → {OUTPUT_FILE}")
    print(f"    Sheet 1: Executive Summary")
    print(f"    Sheet 2: {EXTRACT_SHEET}")


if __name__ == "__main__":
    main()
